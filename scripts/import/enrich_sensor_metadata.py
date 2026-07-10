#!/usr/bin/env python3
"""
Enrich sensor metadata from the Ecosense sensor metadata catalogue.

Aquarius sensor time-series carry hardware/ownership metadata that the API sync
does not populate: the actual instrument model (e.g. SMT100, Implexx Sap Flow
Sensor, FloraPulse_Tensiometer) and the responsible data owner. The DB stores a
generic "Ecosense Node" placeholder in sensor.Sensors.SensorModel instead.

The catalogue keys each series by `external_id` (Aquarius TimeSeriesUniqueID),
which equals our sensor.Sensors.ExternalID. This script matches on that ID and,
for each match:

  - sets SensorModel to the real Instrument (when present)
  - merges Instrument / DataOwner / TypeOfMeasurement / GapTolerance into
    ExternalMetadata (existing keys are preserved unless re-supplied)

Idempotent. Only touches sensors whose ExternalID appears in the catalogue.

By default reads the committed catalogue
data/reference/ecosense_sensor_metadata.csv. A raw Aquarius "Insitu DataUpload"
.xlsx export can be passed instead to refresh from source.

IMPORTANT: run this AFTER an Aquarius sync (sync_aquarius_direct.py). That script
upserts sensors and overwrites sensor_model with the generic "Ecosense Node"
placeholder and resets external_metadata, so this enrichment must be re-applied
each time sensors are re-synced.

    python scripts/import/enrich_sensor_metadata.py [path/to/catalogue.csv|form.xlsx]
"""

import csv
import json
import os
import sys
from pathlib import Path

import psycopg2
from dotenv import load_dotenv
from psycopg2.extras import execute_values

REPO_ROOT = Path(__file__).parent.parent.parent
DEFAULT_CSV = REPO_ROOT / "data" / "reference" / "ecosense_sensor_metadata.csv"
SHEETS = ["SensorData", "deadSensors", "sapflow"]

load_dotenv(REPO_ROOT / "docker" / ".env")

POSTGRES_HOST = "localhost"
POSTGRES_USER = os.getenv("POSTGRES_USER", "postgres")
POSTGRES_PASSWORD = os.getenv("POSTGRES_PASSWORD")
POSTGRES_DATABASE = os.getenv("POSTGRES_DB", "postgres")
POSTGRES_PORT = os.getenv("POSTGRES_PORT", "5432")
POOLER_TENANT_ID = os.getenv("POOLER_TENANT_ID", "")
POSTGRES_USER_POOLER = (
    f"{POSTGRES_USER}.{POOLER_TENANT_ID}" if POOLER_TENANT_ID else POSTGRES_USER
)


def get_db_connection():
    return psycopg2.connect(
        host=POSTGRES_HOST,
        user=POSTGRES_USER_POOLER,
        password=POSTGRES_PASSWORD,
        database=POSTGRES_DATABASE,
        port=POSTGRES_PORT,
    )


def parse_metadata_csv(csv_path):
    """externalid -> {Instrument, DataOwner, TypeOfMeasurement, GapTolerance}."""
    meta = {}
    with open(csv_path, newline="") as f:
        for r in csv.DictReader(f):
            guid = (r.get("external_id") or "").strip()
            if not guid:
                continue
            fields = {
                "Instrument": r.get("instrument"),
                "DataOwner": r.get("data_owner"),
                "TypeOfMeasurement": r.get("measurement_type"),
                "GapTolerance": r.get("gap_tolerance"),
            }
            meta[guid] = {
                k: v.strip() for k, v in fields.items() if v and v.strip()
            }
    print(f"📄 Parsed {len(meta)} sensor metadata entries from {csv_path.name}")
    return meta


def parse_metadata_xlsx(xlsx_path):
    """externalid -> {Instrument, DataOwner, TypeOfMeasurement, GapTolerance}."""
    import openpyxl  # lazy: only needed when refreshing from a raw xlsx export

    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    meta = {}
    for sheet in SHEETS:
        if sheet not in wb.sheetnames:
            continue
        header = None
        for row in wb[sheet].iter_rows(values_only=True):
            if header is None:
                labels = [str(c) if c is not None else "" for c in row]
                if "TimeSeriesUniqueID" in labels:
                    header = {c: i for i, c in enumerate(labels)}
                continue

            def cell(col):
                i = header.get(col)
                return row[i] if i is not None and i < len(row) else None

            guid = cell("TimeSeriesUniqueID")
            if not guid:
                continue
            fields = {
                "Instrument": cell("Instrument"),
                "DataOwner": cell("DataOwner"),
                "TypeOfMeasurement": cell("Type of measurement"),
                "GapTolerance": cell("GapTolerance"),
            }
            # Keep only non-empty values; last sheet wins on collision.
            meta[str(guid).strip()] = {
                k: str(v).strip() for k, v in fields.items() if v not in (None, "")
            }
    print(f"📄 Parsed {len(meta)} sensor metadata entries from {xlsx_path.name}")
    return meta


def enrich(conn, meta):
    cur = conn.cursor()
    cur.execute(
        "SELECT external_id FROM sensor.sensors WHERE external_id IS NOT NULL"
    )
    db_ids = {r[0] for r in cur.fetchall()}
    matched = [eid for eid in db_ids if eid in meta]
    print(f"🔎 {len(db_ids)} sensors with ExternalID; {len(matched)} matched in export")
    if not matched:
        return 0

    rows = []
    for eid in matched:
        m = meta[eid]
        rows.append((eid, m.get("Instrument"), json.dumps(m)))

    # fetch=True aggregates RETURNING rows across all internal pages; cur.rowcount
    # would only reflect the last batch that execute_values sends.
    returned = execute_values(
        cur,
        """
        UPDATE sensor.sensors s SET
            sensor_model = COALESCE(NULLIF(v.instrument, ''), s.sensor_model),
            external_metadata = COALESCE(s.external_metadata, '{}'::jsonb) || v.meta::jsonb,
            updated_by = 'enrich_sensor_metadata_script'
        FROM (VALUES %s) AS v(external_id, instrument, meta)
        WHERE s.external_id = v.external_id
        RETURNING s.sensor_id
        """,
        rows,
        template="(%s, %s, %s::jsonb)",
        fetch=True,
    )
    updated = len(returned)
    conn.commit()
    print(f"✓ Enriched {updated} sensors")
    return updated


def verify(conn):
    cur = conn.cursor()
    cur.execute(
        """
        SELECT sensor_model, COUNT(*)
        FROM sensor.sensors
        WHERE external_metadata ? 'Instrument'
        GROUP BY sensor_model ORDER BY COUNT(*) DESC
        """
    )
    print("\n📊 SensorModel distribution (enriched sensors):")
    for model, count in cur.fetchall():
        print(f"  {model}: {count}")


def main():
    src = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_CSV
    print("=" * 80)
    print("SENSOR METADATA ENRICHMENT")
    print("=" * 80)
    if not src.exists():
        print(f"❌ Metadata source not found: {src}")
        return 1

    conn = get_db_connection()
    try:
        parse = parse_metadata_xlsx if src.suffix.lower() == ".xlsx" else parse_metadata_csv
        meta = parse(src)
        updated = enrich(conn, meta)
        if updated:
            verify(conn)
        print("\n✅ DONE")
        return 0
    except Exception as e:
        conn.rollback()
        print(f"\n❌ Enrichment failed: {e}")
        import traceback

        traceback.print_exc()
        return 1
    finally:
        conn.close()


if __name__ == "__main__":
    sys.exit(main())
